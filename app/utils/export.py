"""Export utilities for Excel and PDF generation."""

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Generate Excel (.xlsx) files from tabular data."""

    def __init__(self, title: str = "Report"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self._row = 1
        self._header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self._header_font = Font(bold=True, color="FFFFFF", size=11)
        self._title_font = Font(bold=True, size=14)

    def add_title(self, title: str, subtitle: str | None = None):
        """Add a title row at the top."""
        self.ws.cell(row=self._row, column=1, value=title).font = self._title_font
        self._row += 1
        if subtitle:
            self.ws.cell(row=self._row, column=1, value=subtitle).font = Font(italic=True, color="666666")
            self._row += 1
        self._row += 1

    def add_headers(self, headers: list[str]):
        """Add header row with styling."""
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self._row, column=col, value=header)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = Alignment(horizontal="center")
        self._row += 1

    def add_row(self, values: list):
        """Add a data row."""
        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=self._row, column=col, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
            elif isinstance(value, date):
                cell.number_format = "YYYY-MM-DD"
        self._row += 1

    def add_rows(self, rows: list[list]):
        """Add multiple data rows."""
        for row in rows:
            self.add_row(row)

    def add_summary_row(self, values: list):
        """Add a bold summary row."""
        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=self._row, column=col, value=value)
            cell.font = Font(bold=True)
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
        self._row += 1

    def auto_width(self):
        """Auto-adjust column widths."""
        for col in range(1, self.ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            self.ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    def to_bytes(self) -> bytes:
        """Return the workbook as bytes."""
        self.auto_width()
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class CSVExporter:
    """Generate CSV from tabular data."""

    def __init__(self):
        self._rows: list[list[str]] = []

    def add_headers(self, headers: list[str]):
        self._rows.append(headers)

    def add_row(self, values: list):
        self._rows.append([str(v) if v is not None else "" for v in values])

    def add_rows(self, rows: list[list]):
        for row in rows:
            self.add_row(row)

    def to_bytes(self) -> bytes:
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerows(self._rows)
        return buffer.getvalue().encode("utf-8-sig")
